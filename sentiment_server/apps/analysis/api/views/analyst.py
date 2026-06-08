import csv
from io import BytesIO, StringIO
from urllib.parse import quote

from django.http import HttpResponse
from django.utils import timezone
from rest_framework import status
from rest_framework.response import Response
from rest_framework.views import APIView

from apps.analysis.application.queries.analyst import (
    build_analyst_comment_list_context,
    build_analyst_overview_context,
    build_analyst_report_context,
)
from apps.analysis.api.responses import StandardResultsSetPagination
from apps.analysis.api.serializers.actions import (
    PublicAnalysisResultSerializer,
)
from apps.analysis.api.serializers.analyst import (
    AnalystCommentFilterSerializer,
    AnalystCommentNoteSerializer,
    AnalystOverviewSerializer,
    AnalystReportExportSerializer,
    AnalystReportSerializer,
)
from apps.analysis.application.commands.analyst_comments import (
    delete_analyst_comment,
    update_analyst_comment,
)
from apps.analysis.infra.selectors.analyst import get_visible_analyst_review_queryset
from apps.users.permissions import IsAnalystOrAdmin
from core.request_ip import get_request_ip


def _serialize_overview_payload(payload):
    return {
        **payload,
        "recent_results": PublicAnalysisResultSerializer(
            payload["recent_results"], many=True
        ).data,
    }


def _serialize_report_payload(payload):
    return payload


def _build_content_disposition(filename, ascii_fallback):
    return (
        f'attachment; filename="{ascii_fallback}"; '
        f"filename*=UTF-8''{quote(filename)}"
    )


def _percent(value):
    return f"{float(value or 0):.1f}%"


def _build_summary_rows(context):
    summary = context["summary"]
    quality = context["quality_summary"]
    range_payload = context["range"]
    return [
        ("统计范围", f"{range_payload['start_date']} 至 {range_payload['end_date']}"),
        ("全部记录", summary["total"]),
        ("积极样本", summary["positive"]),
        ("中性样本", summary["neutral"]),
        ("消极样本", summary["negative"]),
        ("重点关注", summary["priority_count"]),
        ("低置信阈值", _percent(quality["confidence_threshold"] * 100)),
        ("低置信样本", quality["low_confidence_count"]),
        ("低置信占比", _percent(quality["low_confidence_rate"])),
        ("待复核样本", quality["pending_review_count"]),
        ("已审核样本", quality["reviewed_count"]),
        ("审核覆盖率", _percent(quality["review_rate"])),
        ("人工修正样本", quality["corrected_count"]),
        ("已审核修正率", _percent(quality["correction_rate"])),
    ]


def _sentiment_distribution_rows(context):
    labels = [
        ("positive", "积极"),
        ("neutral", "中性"),
        ("negative", "消极"),
    ]
    final_distribution = context["final_sentiment_distribution"]
    model_distribution = context["model_sentiment_distribution"]
    return [
        (display, final_distribution.get(key, 0), model_distribution.get(key, 0))
        for key, display in labels
    ]


def _write_csv_section(writer, title, headers, rows):
    writer.writerow([title])
    writer.writerow(headers)
    writer.writerows(rows)
    writer.writerow([])


def _build_csv_report(context):
    buffer = StringIO()
    writer = csv.writer(buffer)
    _write_csv_section(writer, "摘要", ["指标", "数值"], _build_summary_rows(context))
    _write_csv_section(
        writer,
        "情感分布",
        ["情感", "最终标签数量", "模型原始数量"],
        _sentiment_distribution_rows(context),
    )
    _write_csv_section(
        writer,
        "趋势明细",
        ["日期", "总数", "积极", "中性", "消极", "积极率"],
        [
            (
                row["date"],
                row["total"],
                row["positive"],
                row["neutral"],
                row["negative"],
                _percent(row["positive_rate"]),
            )
            for row in context["detail_rows"]
        ],
    )
    _write_csv_section(
        writer,
        "分类分布",
        ["分类", "数量"],
        [(row["category"], row["count"]) for row in context["category_distribution"]],
    )
    _write_csv_section(
        writer,
        "项目分布",
        ["项目", "数量"],
        [(row["label"], row["value"]) for row in context["project_distribution"]],
    )
    _write_csv_section(
        writer,
        "来源分布",
        ["来源", "数量"],
        [(row["label"], row["value"]) for row in context["source_distribution"]],
    )
    _write_csv_section(
        writer,
        "置信度分布",
        ["区间", "数量"],
        [(row["label"], row["value"]) for row in context["confidence_buckets"]],
    )
    _write_csv_section(
        writer,
        "高频关键词",
        ["关键词", "次数"],
        [(row["keyword"], row["count"]) for row in context["keyword_top"]],
    )
    _write_csv_section(
        writer,
        "模型与最终标签差异",
        ["模型标签", "最终标签", "数量"],
        [
            (
                row["model_sentiment_display"],
                row["final_sentiment_display"],
                row["count"],
            )
            for row in context["correction_matrix"]
        ],
    )
    return ("\ufeff" + buffer.getvalue()).encode("utf-8")


def _append_sheet(workbook, title, headers, rows):
    worksheet = workbook.create_sheet(title=title)
    worksheet.append(headers)
    for row in rows:
        worksheet.append(list(row))
    return worksheet


def _build_xlsx_report(context):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill

    workbook = Workbook()
    workbook.remove(workbook.active)
    sheets = [
        ("摘要", ["指标", "数值"], _build_summary_rows(context)),
        (
            "情感分布",
            ["情感", "最终标签数量", "模型原始数量"],
            _sentiment_distribution_rows(context),
        ),
        (
            "趋势明细",
            ["日期", "总数", "积极", "中性", "消极", "积极率"],
            [
                (
                    row["date"],
                    row["total"],
                    row["positive"],
                    row["neutral"],
                    row["negative"],
                    _percent(row["positive_rate"]),
                )
                for row in context["detail_rows"]
            ],
        ),
        (
            "分类分布",
            ["分类", "数量"],
            [(row["category"], row["count"]) for row in context["category_distribution"]],
        ),
        (
            "项目分布",
            ["项目", "数量"],
            [(row["label"], row["value"]) for row in context["project_distribution"]],
        ),
        (
            "来源分布",
            ["来源", "数量"],
            [(row["label"], row["value"]) for row in context["source_distribution"]],
        ),
        (
            "置信度分布",
            ["区间", "数量"],
            [(row["label"], row["value"]) for row in context["confidence_buckets"]],
        ),
        (
            "高频关键词",
            ["关键词", "次数"],
            [(row["keyword"], row["count"]) for row in context["keyword_top"]],
        ),
        (
            "标签差异",
            ["模型标签", "最终标签", "数量"],
            [
                (
                    row["model_sentiment_display"],
                    row["final_sentiment_display"],
                    row["count"],
                )
                for row in context["correction_matrix"]
            ],
        ),
    ]

    for title, headers, rows in sheets:
        worksheet = _append_sheet(workbook, title, headers, rows)
        for cell in worksheet[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor="4F46E5")
        for column in worksheet.columns:
            max_length = max(len(str(cell.value or "")) for cell in column)
            worksheet.column_dimensions[column[0].column_letter].width = min(
                max(max_length + 2, 12),
                42,
            )

    stream = BytesIO()
    workbook.save(stream)
    workbook.close()
    return stream.getvalue()


def _build_report_export_response(context, export_format):
    timestamp = timezone.localtime().strftime("%Y%m%d_%H%M%S")
    if export_format == "csv":
        filename = f"分析师报表_{timestamp}.csv"
        response = HttpResponse(
            _build_csv_report(context),
            content_type="text/csv; charset=utf-8",
        )
        ascii_fallback = f"analyst-report-{timestamp}.csv"
    else:
        filename = f"分析师报表_{timestamp}.xlsx"
        response = HttpResponse(
            _build_xlsx_report(context),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        ascii_fallback = f"analyst-report-{timestamp}.xlsx"

    response["Content-Disposition"] = _build_content_disposition(
        filename,
        ascii_fallback,
    )
    return response


class AnalystOverviewView(APIView):
    permission_classes = [IsAnalystOrAdmin]

    def get(self, request):
        serializer = AnalystOverviewSerializer(data=request.query_params)
        if not serializer.is_valid():
            return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

        context = build_analyst_overview_context(
            user=request.user,
            validated_data=serializer.validated_data,
        )
        return Response(_serialize_overview_payload(context))


class AnalystCommentListView(APIView):
    permission_classes = [IsAnalystOrAdmin]

    def get(self, request):
        serializer = AnalystCommentFilterSerializer(data=request.query_params)
        if not serializer.is_valid():
            return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

        context = build_analyst_comment_list_context(
            user=request.user,
            validated_data=serializer.validated_data,
        )
        paginator = StandardResultsSetPagination()
        paginator.page_size = serializer.validated_data["page_size"]
        page = paginator.paginate_queryset(context["queryset"], request, view=self)
        response = paginator.get_paginated_response(
            PublicAnalysisResultSerializer(page, many=True).data
        )
        response.data["category_options"] = context["category_options"]
        return Response(response.data)


class AnalystCommentDetailView(APIView):
    permission_classes = [IsAnalystOrAdmin]

    def get_object(self, pk):
        return get_visible_analyst_review_queryset(self.request.user).filter(pk=pk).first()

    def patch(self, request, pk):
        result = self.get_object(pk)
        if not result:
            return Response(
                {"error": "分析结果不存在"}, status=status.HTTP_404_NOT_FOUND
            )

        serializer = AnalystCommentNoteSerializer(
            result, data=request.data, partial=True
        )
        if not serializer.is_valid():
            return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

        result = update_analyst_comment(
            result=result,
            validated_data=serializer.validated_data,
            user=request.user,
            client_ip=get_request_ip(request),
        )
        return Response(PublicAnalysisResultSerializer(result).data)

    def delete(self, request, pk):
        result = self.get_object(pk)
        if not result:
            return Response(
                {"error": "分析结果不存在"}, status=status.HTTP_404_NOT_FOUND
            )

        delete_analyst_comment(
            result=result,
            user=request.user,
            client_ip=get_request_ip(request),
        )
        return Response(status=status.HTTP_204_NO_CONTENT)


class AnalystReportView(APIView):
    permission_classes = [IsAnalystOrAdmin]

    def get(self, request):
        serializer = AnalystReportSerializer(data=request.query_params)
        if not serializer.is_valid():
            return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

        context = build_analyst_report_context(
            user=request.user,
            validated_data=serializer.validated_data,
        )
        return Response(_serialize_report_payload(context))


class AnalystReportExportView(APIView):
    permission_classes = [IsAnalystOrAdmin]

    def get(self, request):
        serializer = AnalystReportExportSerializer(data=request.query_params)
        if not serializer.is_valid():
            return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

        validated_data = dict(serializer.validated_data)
        export_format = validated_data.pop("format", "xlsx")
        context = build_analyst_report_context(
            user=request.user,
            validated_data=validated_data,
        )
        return _build_report_export_response(context, export_format)
